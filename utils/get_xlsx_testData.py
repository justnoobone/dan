# -*- coding: utf-8 -*-
# 获取xlsx文件
import pytest
from openpyxl import load_workbook


def get_xlsx_test_data(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    data = []
    # 将生成器对象转换为列表，然后从第二行开始读取数据
    for row in list(sheet.iter_rows(values_only=True))[1:]:
        data.append(row)
    return data


@pytest.mark.parametrize('name,tele,password', get_xlsx_test_data('./testData/import.xlsx'))
def test_example(name, tele, password):
    print("姓名：", name,end=' ')
    print("手机号码：", tele,end=' ')
    print("手机型号：", password,end=' ')


if __name__ == '__main__':
    pytest.main([__file__])